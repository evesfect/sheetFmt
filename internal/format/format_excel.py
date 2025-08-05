#!/usr/bin/env python3
import json
import sys
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import logging
from datetime import datetime



def clean_column_name(raw_name: str) -> str:
    """Clean column names by removing HTML tags, extra whitespace, and taking first line"""
    if not raw_name:
        return ""
    
    # Convert to string and strip basic whitespace
    cleaned = str(raw_name).strip()
    
    # Remove HTML/XML tags using regex
    # This handles both self-closing and regular tags
    cleaned = re.sub(r'<[^>]+>', '', cleaned)
    
    # Split by newlines and take the first non-empty line
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    if lines:
        cleaned = lines[0]
    else:
        cleaned = ""
    
    # Remove extra whitespace and normalize
    cleaned = ' '.join(cleaned.split())
    
    return cleaned

def setup_logging():
    """Setup logging configuration with Unicode support"""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    # Create file handler with UTF-8 encoding
    file_handler = logging.FileHandler(
        log_dir / "format_excel.log", 
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    
    # Create console handler 
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    
    # Create formatters
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_formatter = logging.Formatter('%(levelname)s - %(message)s')
    
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(console_formatter)
    
    # Setup logger
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


logger = setup_logging()


def safe_print(text):
    """Print text with encoding error handling"""
    try:
        print(text)
    except UnicodeEncodeError:
        safe_text = text.encode('ascii', 'replace').decode('ascii')
        print(safe_text)


class ExcelFormatter:
    def __init__(self, input_file: str, target_file: str, mapping_file: str, 
                 output_file: str, input_sheet: str, target_sheet: str, 
                 table_end_tolerance: int = 1, clean_formula_only_rows: bool = True):
        self.input_file = input_file
        self.target_file = target_file
        self.mapping_file = mapping_file
        self.output_file = output_file
        self.input_sheet_name = input_sheet
        self.target_sheet_name = target_sheet
        self.table_end_tolerance = table_end_tolerance
        self.clean_formula_only_rows = clean_formula_only_rows
        
        self.input_wb = None
        self.target_wb = None
        self.input_ws = None
        self.target_ws = None
        
        self.mapping_config = None
        self.scanned_to_target = {}  # Changed: scanned -> target
        self.ignored_scanned = set()  # Changed: track ignored scanned columns
        self.error_messages = []
        self.sections = []
        self._necessary_columns = None
        self._current_applicable_mappings = {}
        
    def load_files(self):
        """Load all required files"""
        try:
            # Check if mapping file exists
            if not os.path.exists(self.mapping_file):
                raise Exception(f"Mapping file not found: {self.mapping_file}")
            
            # Load mapping configuration with UTF-8 encoding
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                self.mapping_config = json.load(f)
            
            logger.debug(f"Loaded mapping file: {self.mapping_file}")
            logger.debug(f"Mapping config keys: {list(self.mapping_config.keys())}")
            
            # Create scanned -> target mapping and ignored set
            mappings_count = 0
            ignored_count = 0
            
            for mapping in self.mapping_config.get('mappings', []):
                scanned_col = mapping['scanned_column']
                
                if mapping.get('is_ignored', False):
                    # Track ignored scanned columns
                    self.ignored_scanned.add(scanned_col)
                    ignored_count += 1
                    logger.debug(f"Ignoring scanned column '{scanned_col}'")
                elif mapping.get('target_column'):
                    target_col = mapping['target_column']
                    # Changed: Now we map scanned -> target (can have multiple scanned for same target)
                    self.scanned_to_target[scanned_col] = target_col
                    mappings_count += 1
                    logger.debug(f"Mapping scanned '{scanned_col}' -> target '{target_col}'")
            
            logger.debug(f"Total mappings loaded: {mappings_count}")
            logger.debug(f"Total ignored: {ignored_count}")
            
            # Load Excel files
            # CHANGED: Load input file with data_only=True to get calculated values, not formulas
            self.input_wb = load_workbook(self.input_file, data_only=True)
            # Keep target file with data_only=False to preserve its structure
            self.target_wb = load_workbook(self.target_file, data_only=False)
            
            # Get worksheets
            logger.debug(f"Input sheets: {self.input_wb.sheetnames}")
            logger.debug(f"Target sheets: {self.target_wb.sheetnames}")
            
            self.input_ws = self.input_wb[self.input_sheet_name]
            self.target_ws = self.target_wb[self.target_sheet_name]
            
        except Exception as e:
            logger.error(f"Failed to load files: {e}")
            raise Exception(f"Failed to load files: {e}")

    def load_necessary_columns(self) -> List[str]:
        """Load necessary column names from the necessary_columns file"""
        necessary_columns_file = "data/output/necessary_columns"
        
        if not os.path.exists(necessary_columns_file):
            logger.info("No necessary_columns file found, skipping necessary column validation")
            return []
        
        try:
            with open(necessary_columns_file, 'r', encoding='utf-8') as f:
                necessary_columns = [line.strip() for line in f if line.strip()]
            
            logger.info(f"Loaded {len(necessary_columns)} necessary columns from file")
            
            # Validate that necessary columns exist in target columns
            target_columns = self.load_target_columns_from_file()
            target_columns_set = {col.lower().strip() for col in target_columns}
            
            missing_columns = []
            for col in necessary_columns:
                if col.lower().strip() not in target_columns_set:
                    missing_columns.append(col)
            
            if missing_columns:
                error_msg = f"Necessary columns not found in target_columns: {missing_columns}"
                logger.error(error_msg)
                raise Exception(error_msg)
            
            logger.debug(f"All necessary columns validated against target_columns: {necessary_columns}")
            return necessary_columns
            
        except Exception as e:
            if "not found in target_columns" in str(e):
                raise e
            logger.error(f"Failed to load necessary columns file: {e}")
            raise Exception(f"Failed to load necessary columns file: {e}")

    def check_row_has_necessary_data(self, input_row_idx: int, input_headers: Dict[str, int], 
                                applicable_mappings: Dict[str, str], necessary_columns: List[str]) -> bool:
        """Check if a row has data in all necessary columns"""
        if not necessary_columns:
            return True  # No necessary columns defined, all rows are valid
        
        for necessary_col in necessary_columns:
            # Check if this necessary column is mapped and has data
            if necessary_col in applicable_mappings:
                input_column = applicable_mappings[necessary_col]
                if input_column in input_headers:
                    input_col_idx = input_headers[input_column]
                    cell = self.input_ws.cell(row=input_row_idx, column=input_col_idx)
                    
                    # Check if cell has meaningful data
                    if cell.value is None or str(cell.value).strip() == "":
                        logger.debug(f"Row {input_row_idx} missing necessary data in column '{necessary_col}' (mapped from '{input_column}')")
                        return False
                else:
                    logger.debug(f"Row {input_row_idx} missing necessary column '{necessary_col}' - input column '{input_column}' not found")
                    return False
            else:
                logger.debug(f"Row {input_row_idx} missing necessary column '{necessary_col}' - no mapping found")
                return False
        
        return True

    def check_and_remove_empty_files(self, output_files: List[str]) -> List[str]:
        """Check output files for actual data and remove empty ones"""
        logger.info("Checking for empty output files")
        
        valid_files = []
        removed_files = []
        
        for output_file in output_files:
            if self.file_has_data(output_file):
                valid_files.append(output_file)
                logger.debug(f"File has data, keeping: {os.path.basename(output_file)}")
            else:
                try:
                    os.remove(output_file)
                    removed_files.append(output_file)
                    logger.info(f"Removed empty file: {os.path.basename(output_file)}")
                except Exception as e:
                    logger.error(f"Failed to remove empty file {output_file}: {e}")
                    # Keep in valid_files if we can't remove it
                    valid_files.append(output_file)
        
        if removed_files:
            logger.info(f"Removed {len(removed_files)} empty files")
            safe_print(f"🗑️  Removed {len(removed_files)} empty files:")
            for removed_file in removed_files:
                safe_print(f"  - {os.path.basename(removed_file)}")
        
        return valid_files

    def file_has_data(self, file_path: str) -> bool:
        """Check if an Excel file has actual data rows (beyond header)"""
        try:
            # Open the file to check for data
            check_wb = load_workbook(file_path, read_only=True, data_only=True)
            check_ws = check_wb.active
            
            # Find header row
            header_row = self.detect_header_row(check_ws)
            data_start_row = header_row + 1
            
            # Check if there are any rows with data after header
            has_data = False
            for row_idx in range(data_start_row, check_ws.max_row + 1):
                for col_idx in range(1, check_ws.max_column + 1):
                    cell = check_ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != "":
                        has_data = True
                        break
                if has_data:
                    break
            
            check_wb.close()
            
            logger.debug(f"File {os.path.basename(file_path)} has data: {has_data}")
            return has_data
            
        except Exception as e:
            logger.error(f"Error checking file for data {file_path}: {e}")
            return True  # If we can't check, assume it has data to be safe

    def load_target_columns_from_file(self) -> List[str]:
        """Load target column names from the target_columns file"""
        # Try to find target_columns file in the expected location
        target_columns_file = "data/output/target_columns"
        
        if not os.path.exists(target_columns_file):
            logger.warning(f"Target columns file not found at {target_columns_file}, using mapping file columns")
            # Fallback to target columns from mapping config
            target_columns = set()
            for mapping in self.mapping_config.get('mappings', []):
                if mapping.get('target_column') and not mapping.get('is_ignored', False):
                    target_columns.add(mapping['target_column'])
            return list(target_columns)
        
        try:
            with open(target_columns_file, 'r', encoding='utf-8') as f:
                target_columns = [line.strip() for line in f if line.strip()]
            
            logger.debug(f"Loaded {len(target_columns)} target columns from file")
            return target_columns
        except Exception as e:
            logger.error(f"Failed to load target columns file: {e}")
            return []

    def detect_header_rows_by_target_match(self) -> List[int]:
        """Detect header rows by checking if any cell matches target column names"""
        logger.debug("Detecting header rows by matching target columns")
        
        target_columns = self.load_target_columns_from_file()
        if not target_columns:
            logger.warning("No target columns loaded, using original header detection")
            return [self.detect_header_row(self.input_ws)]
        
        # Create set for faster lookup (case-insensitive)
        target_columns_set = {col.lower().strip() for col in target_columns}
        logger.debug(f"Checking against {len(target_columns_set)} target columns")
        
        header_rows = []
        
        # Check each row in the input worksheet
        for row_idx in range(1, self.input_ws.max_row + 1):
            row_has_header = False
            
            # Check each cell in the row
            for col_idx in range(1, self.input_ws.max_column + 1):
                cell = self.input_ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    
                    # Clean the cell value same way we clean column names
                    cleaned_cell_value = self.clean_column_name_for_matching(str(cell.value))
                    
                    if cleaned_cell_value.lower() in target_columns_set:
                        logger.debug(f"Found target column match at row {row_idx}, col {col_idx}: '{cleaned_cell_value}'")
                        row_has_header = True
                        break
            
            if row_has_header:
                header_rows.append(row_idx)
                logger.debug(f"Marked row {row_idx} as header row")
        
        if not header_rows:
            logger.warning("No header rows found by target matching, falling back to original detection")
            return [self.detect_header_row(self.input_ws)]
        
        logger.info(f"Found {len(header_rows)} header rows by target matching: {header_rows}")
        return header_rows

    def clean_column_name_for_matching(self, raw_name: str) -> str:
        """Clean column name for matching (same logic as in scanner.go)"""
        if not raw_name:
            return ""
        
        # Convert to string and strip basic whitespace
        cleaned = str(raw_name).strip()
        
        # Remove HTML/XML tags using regex
        import re
        cleaned = re.sub(r'<[^>]+>', '', cleaned)
        
        # Split by newlines and take the first non-empty line
        lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
        if lines:
            cleaned = lines[0]
        else:
            cleaned = ""
        
        # Remove extra whitespace and normalize
        cleaned = ' '.join(cleaned.split())
        
        return cleaned

    def split_input_sheet_by_headers(self, header_rows: List[int]) -> List[Tuple[int, int, int]]:
        """Split input sheet into sections based on header rows. Returns list of (header_row, start_row, end_row)"""
        if len(header_rows) <= 1:
            # Single section
            header_row = header_rows[0] if header_rows else 1
            start_row = header_row + 1
            end_row = self.input_ws.max_row
            return [(header_row, start_row, end_row)]
        
        sections = []
        
        for i, header_row in enumerate(header_rows):
            start_row = header_row + 1
            
            # Determine end row
            if i < len(header_rows) - 1:
                end_row = header_rows[i + 1] - 1
            else:
                end_row = self.input_ws.max_row
            
            # Only add section if it has data rows
            if start_row <= end_row:
                sections.append((header_row, start_row, end_row))
                logger.debug(f"Section {i+1}: header row {header_row}, data rows {start_row}-{end_row}")
        
        logger.info(f"Split input sheet into {len(sections)} sections")
        return sections

    def process_sheet_section(self, section_num: int, header_row: int, start_row: int, end_row: int) -> str:
        """Process a single section of the input sheet and return the output filename"""
        logger.info(f"Processing section {section_num}: header row {header_row}, data rows {start_row}-{end_row}")
        
        # Create output filename for this section
        base_name = os.path.splitext(os.path.basename(self.output_file))[0]
        if len(self.sections) > 1:  # Only add part number if multiple sections
            section_output = f"{base_name}_part{section_num}.xlsx"
        else:
            section_output = f"{base_name}.xlsx"
        
        section_output_path = os.path.join(os.path.dirname(self.output_file), section_output)
        
        # Create a new target workbook for this section
        section_target_wb = load_workbook(self.target_file, data_only=False)
        section_target_ws = section_target_wb[self.target_sheet_name]
        
        # Store original values
        original_input_header_row = getattr(self, 'input_header_row', 1)
        original_target_wb = self.target_wb
        original_target_ws = self.target_ws
        original_output_file = self.output_file
        
        # Temporarily update for this section
        self.input_header_row = header_row
        self.target_wb = section_target_wb
        self.target_ws = section_target_ws
        self.output_file = section_output_path
        
        try:
            # Get headers for this section
            section_input_headers = self.get_section_input_headers(header_row)
            target_headers = self.get_target_headers()
            
            # Find applicable mappings for this section
            applicable_mappings = self.find_applicable_mappings(section_input_headers, target_headers)
            # Store applicable mappings for necessary column validation
            self._current_applicable_mappings = applicable_mappings
            
            # Calculate section data range
            section_data_rows = end_row - start_row + 1
            target_max_needed_row = self.target_header_row + section_data_rows
            max_target_row = max(self.target_ws.max_row, target_max_needed_row)
            
            # Clear target data
            for header_name, col_idx in target_headers:
                if header_name:
                    self.clear_column_data(col_idx, max_target_row)
            
            # Process each target column
            for header_name, col_idx in target_headers:
                if not header_name:
                    continue
                    
                if header_name in applicable_mappings:
                    input_column = applicable_mappings[header_name]
                    self.process_section_column_mapping(
                        col_idx, header_name, input_column, section_input_headers, 
                        start_row, end_row
                    )
            
            # Clean formula-only rows
            self.clean_formula_only_rows_func()
            
            # Save this section
            section_target_wb.save(section_output_path)
            logger.info(f"Section {section_num} saved to: {section_output_path}")
            
            return section_output_path
            
        finally:
            # Restore original values
            self.input_header_row = original_input_header_row
            self.target_wb = original_target_wb
            self.target_ws = original_target_ws
            self.output_file = original_output_file
            section_target_wb.close()

    def get_section_input_headers(self, header_row: int) -> Dict[str, int]:
        """Get input headers for a specific section"""
        headers = {}
        
        logger.debug(f"Getting section headers from row {header_row}")
        
        for col_idx in range(1, self.input_ws.max_column + 1):
            cell = self.input_ws.cell(row=header_row, column=col_idx)
            if cell.value:
                raw_header = str(cell.value).strip()
                clean_header = self.clean_column_name_for_matching(raw_header)
                
                if clean_header:
                    headers[clean_header] = col_idx
                    logger.debug(f"Section header - Column {col_idx}: '{clean_header}'")
        
        return headers

    def process_section_column_mapping(self, target_col_idx: int, target_header: str, 
                                 input_column: str, input_headers: Dict[str, int],
                                 data_start_row: int, data_end_row: int):
        """Process column mapping for a specific section with necessary column validation"""
        logger.debug(f"Processing section column '{target_header}' mapped from '{input_column}'")
        
        input_col_idx = input_headers[input_column]
        target_data_start_row = self.target_header_row + 1
        
        # Load necessary columns once for this section
        necessary_columns = getattr(self, '_necessary_columns', None)
        if necessary_columns is None:
            self._necessary_columns = self.load_necessary_columns()
            necessary_columns = self._necessary_columns
        
        # Get applicable mappings for necessary column validation
        applicable_mappings = getattr(self, '_current_applicable_mappings', {})
        
        target_row = target_data_start_row
        skipped_rows = 0
        processed_rows = 0
        
        for input_row_idx in range(data_start_row, data_end_row + 1):
            # Check if this row has all necessary data
            if not self.check_row_has_necessary_data(input_row_idx, input_headers, applicable_mappings, necessary_columns):
                skipped_rows += 1
                continue
            
            input_cell = self.input_ws.cell(row=input_row_idx, column=input_col_idx)
            target_cell = self.target_ws.cell(row=target_row, column=target_col_idx)
            
            # Skip formulas
            if isinstance(input_cell.value, str) and input_cell.value.startswith('='):
                continue
                
            self.copy_cell_value_with_type_preservation(input_cell, target_cell)
            target_row += 1
            processed_rows += 1
        
        if skipped_rows > 0:
            logger.info(f"Column '{target_header}': processed {processed_rows} rows, skipped {skipped_rows} rows due to missing necessary data")
        else:
            logger.debug(f"Column '{target_header}': processed {processed_rows} rows")

    def detect_table_end_row(self, worksheet, header_row: int) -> int:
        """Detect where the data table ends by checking rightmost column continuity with tolerance"""
        
        if worksheet.max_row == 0:
            return header_row  # No data, table ends at header
        
        # Find the rightmost column with data (same logic as header detection)
        rightmost_col = 0
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(worksheet.max_column, 0, -1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    if col_idx > rightmost_col:
                        rightmost_col = col_idx
                    break
        
        if rightmost_col == 0:
            return header_row  # No data found
        
        logger.debug(f"Using table end tolerance: {self.table_end_tolerance}")
        
        # Starting from header row, find where table ends
        table_end_row = header_row
        
        for row_idx in range(header_row, worksheet.max_row + 1):
            current_cell = worksheet.cell(row=row_idx, column=rightmost_col)
            current_has_data = current_cell.value is not None and str(current_cell.value).strip() != ""
            
            if current_has_data:
                # Check if next N rows (tolerance) have data in rightmost column
                empty_rows_count = 0
                
                for check_offset in range(1, self.table_end_tolerance + 1):
                    check_row_idx = row_idx + check_offset
                    if check_row_idx <= worksheet.max_row:
                        check_cell = worksheet.cell(row=check_row_idx, column=rightmost_col)
                        check_has_data = check_cell.value is not None and str(check_cell.value).strip() != ""
                        
                        if not check_has_data:
                            empty_rows_count += 1
                        else:
                            break  # Found data within tolerance, continue with main loop
                    else:
                        # Beyond worksheet bounds, count as empty
                        empty_rows_count += 1
                
                # If all tolerance rows are empty, current row is table end
                if empty_rows_count == self.table_end_tolerance:
                    table_end_row = row_idx
                    break
                else:
                    # Continue checking, update table end to current row
                    table_end_row = row_idx
        
        logger.debug(f"Table end row detected: {table_end_row} (tolerance: {self.table_end_tolerance})")
        return table_end_row

    def clean_formula_only_rows_func(self):
        """Remove rows that only contain formulas and completely empty rows"""
        if not self.clean_formula_only_rows:
            logger.debug("Formula-only row cleaning is disabled")
            return
        
        logger.debug("Starting formula-only and empty row cleanup...")
        
        rows_to_delete = []
        
        # Check each row starting from row 1, but skip header row
        for row_idx in range(1, self.target_ws.max_row + 1):
            # Skip the header row
            if row_idx == self.target_header_row:
                continue
                
            has_non_formula_data = False
            
            # Check all cells in this row
            for col_idx in range(1, self.target_ws.max_column + 1):
                cell = self.target_ws.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None:
                    # If cell contains data that's not a formula, keep the row
                    if cell.data_type != 'f':
                        # Check if it's not just whitespace
                        if str(cell.value).strip() != "":
                            has_non_formula_data = True
                            break
            
            # If row has no non-formula data (either empty or only formulas), mark for deletion
            if not has_non_formula_data:
                rows_to_delete.append(row_idx)
        
        # Delete rows in reverse order to maintain row indices
        deleted_count = 0
        for row_idx in reversed(rows_to_delete):
            self.target_ws.delete_rows(row_idx, 1)
            deleted_count += 1
        
        logger.debug(f"Cleaned up {deleted_count} empty/formula-only rows")

    def get_input_headers(self) -> Dict[str, int]:
        """Get input file headers mapping"""
        headers = {}
        
        # Detect header row
        input_header_row = self.detect_header_row(self.input_ws)
        logger.debug(f"Input header row detected: {input_header_row}")
        
        # Detect table end row
        self.input_table_end_row = self.detect_table_end_row(self.input_ws, input_header_row)
        logger.debug(f"Input table end row detected: {self.input_table_end_row}")
        
        logger.debug("Input file headers:")
        for col_idx in range(1, self.input_ws.max_column + 1):
            cell = self.input_ws.cell(row=input_header_row, column=col_idx)
            if cell.value:
                raw_header = str(cell.value).strip()
                # NEW: Clean the header name
                clean_header = clean_column_name(raw_header)
                
                if clean_header:  # Only add if we have a clean name
                    headers[clean_header] = col_idx
                    
                    # Log both raw and cleaned versions if they're different
                    if raw_header != clean_header:
                        logger.debug(f"  Column {col_idx}: '{clean_header}' (cleaned from: '{raw_header[:50]}...')")
                    else:
                        logger.debug(f"  Column {col_idx}: '{clean_header}'")
                else:
                    logger.debug(f"  Column {col_idx}: [EMPTY after cleaning] (raw: '{raw_header[:50]}...')")
        
        # Store the header row for later use
        self.input_header_row = input_header_row
        return headers
    
    def get_target_headers(self) -> List[Tuple[str, int]]:
        """Get target format headers with their column indices"""
        headers = []
        
        # Detect header row
        target_header_row = self.detect_header_row(self.target_ws)
        logger.debug(f"Target header row detected: {target_header_row}")
        
        logger.debug("Target format headers:")
        for col_idx in range(1, self.target_ws.max_column + 1):
            cell = self.target_ws.cell(row=target_header_row, column=col_idx)
            header_value = str(cell.value).strip() if cell.value else ""
            headers.append((header_value, col_idx))
            logger.debug(f"  Column {col_idx}: '{header_value}'")
        
        # Store the header row for later use
        self.target_header_row = target_header_row
        return headers
    
    def find_applicable_mappings(self, input_headers: Dict[str, int], target_headers: List[Tuple[str, int]]) -> Dict[str, str]:
        """Find which input columns can be mapped to which target columns based on available data"""
        applicable_mappings = {}  # target_column -> input_column
        
        logger.debug("=== FINDING APPLICABLE MAPPINGS ===")
        logger.debug(f"Available input columns: {list(input_headers.keys())}")
        
        # Get all target column names
        target_column_names = [header for header, _ in target_headers if header]
        logger.debug(f"Target columns needing data: {target_column_names}")
        
        # For each target column, find if any available input column maps to it
        for target_col in target_column_names:
            found_mapping = False
            
            # Look through all available input columns
            for input_col in input_headers.keys():
                # Check if this input column is ignored
                if input_col in self.ignored_scanned:
                    logger.debug(f"  Skipping ignored input column '{input_col}'")
                    continue
                
                # Check if this input column has a mapping to our target column
                if input_col in self.scanned_to_target and self.scanned_to_target[input_col] == target_col:
                    applicable_mappings[target_col] = input_col
                    logger.debug(f"  ✓ APPLICABLE: Target '{target_col}' <- Input '{input_col}' (found in file)")
                    found_mapping = True
                    break  # Use first available mapping for this target
            
            if not found_mapping:
                logger.debug(f"  ✗ NO MAPPING: Target '{target_col}' - no available input column maps to it")
        
        logger.debug(f"=== FINAL APPLICABLE MAPPINGS: {len(applicable_mappings)} ===")
        for target_col, input_col in applicable_mappings.items():
            logger.debug(f"  '{target_col}' <- '{input_col}'")
        
        return applicable_mappings
    
    

    def copy_cell_value_with_type_preservation(self, source_cell, target_cell):
        """Copy cell value (never formulas) and data type from source to target"""
        
        # Get the actual value (never a formula)
        source_value = source_cell.value
        
        # Extra safety check: if somehow a formula string got through, don't copy it
        if isinstance(source_value, str) and source_value.startswith('='):
            logger.warning(f"Detected formula string in source cell, skipping: {source_value[:50]}...")
            return  # Don't copy formulas
        
        # Copy the calculated value with type preservation
        if source_cell.data_type == 'd':  # Date
            target_cell.value = source_value
        elif source_cell.data_type == 'n':  # Number
            target_cell.value = source_value
        elif source_cell.data_type == 'b':  # Boolean
            target_cell.value = source_value
        else:  # String or other (but not formula since we loaded with data_only=True)
            target_cell.value = source_value
        
        # Copy number format to preserve data type appearance (but not if it's a formula format)
        if source_cell.number_format and source_cell.number_format != 'General':
            target_cell.number_format = source_cell.number_format
        
        # Log what we're copying for debugging
        if source_value is not None:
            logger.debug(f"Copied value: {source_value} (type: {source_cell.data_type})")
    
    def process_column_with_mapping(self, target_col_idx: int, target_header: str, 
                           input_column: str, input_headers: Dict[str, int]):
        """Process a column that has data mapping"""
        logger.debug(f"Processing column '{target_header}' mapped from '{input_column}'")
        
        input_col_idx = input_headers[input_column]
        logger.debug(f"Input column '{input_column}' is at index {input_col_idx}")
        
        # Get the maximum row with data in input file
        max_input_row = self.input_ws.max_row
        
        # Process each data row from input file (starting after header row)
        rows_processed = 0
        cells_copied = 0
        formulas_skipped = 0  # Track skipped formulas
        input_data_start_row = self.input_header_row + 1
        target_data_start_row = self.target_header_row + 1
        
        logger.debug(f"Processing rows from input row {input_data_start_row} to {max_input_row}")
        logger.debug(f"Mapping to target starting at row {target_data_start_row}")
        
        for input_row_idx in range(input_data_start_row, max_input_row + 1):
            # Calculate corresponding target row
            target_row_idx = target_data_start_row + (input_row_idx - input_data_start_row)
            
            input_cell = self.input_ws.cell(row=input_row_idx, column=input_col_idx)
            target_cell = self.target_ws.cell(row=target_row_idx, column=target_col_idx)
            
            # Additional safety check for formulas
            input_value = input_cell.value
            if isinstance(input_value, str) and input_value.startswith('='):
                logger.warning(f"Skipping formula at input[{input_row_idx},{input_col_idx}]: {input_value[:30]}...")
                formulas_skipped += 1
                continue
            
            # Log what we're copying (only non-empty, non-formula values)
            if input_value is not None and str(input_value).strip() != "":
                logger.debug(f"Copying from input[{input_row_idx},{input_col_idx}] = '{input_value}' to target[{target_row_idx},{target_col_idx}]")
                cells_copied += 1
            
            # Copy data (this method now has additional formula protection)
            self.copy_cell_value_with_type_preservation(input_cell, target_cell)
            rows_processed += 1
        
        if formulas_skipped > 0:
            logger.warning(f"Skipped {formulas_skipped} formula cells in column '{target_header}'")
        
        logger.debug(f"Processed {rows_processed} rows for column '{target_header}', copied {cells_copied} non-empty cells")

    def clear_column_data(self, col_idx: int, max_row: int):
        """Clear data in a column"""
        cleared_cells = 0
        target_data_start_row = self.target_header_row + 1
        
        for row_idx in range(target_data_start_row, max_row + 1):
            cell = self.target_ws.cell(row=row_idx, column=col_idx)
            cell.value = None
            cleared_cells += 1
        
        if cleared_cells > 0:
            col_letter = get_column_letter(col_idx)
            logger.debug(f"Cleared {cleared_cells} cells in column {col_letter}")

    def detect_header_row(self, worksheet) -> int:
        """Detect header row using the strategy: find rightmost column with data, 
        then find first row with data in that column"""
        
        if worksheet.max_row == 0:
            return 1  # Default to row 1 if no data
        
        # Find the rightmost column with data across all rows
        rightmost_col = 0
        for row_idx in range(1, worksheet.max_row + 1):
            for col_idx in range(worksheet.max_column, 0, -1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    if col_idx > rightmost_col:
                        rightmost_col = col_idx
                    break  # Found the rightmost data in this row
        
        if rightmost_col == 0:
            return 1  # Default to row 1 if no data
        
        # Now find the first row that has data in the rightmost column
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=rightmost_col)
            if cell.value is not None and str(cell.value).strip() != "":
                return row_idx
        
        return 1  # Default to row 1 if not found

    def format_excel(self):
        """Main formatting function with section splitting support"""
        try:
            logger.info(f"Starting Excel formatting for {os.path.basename(self.input_file)}")
            self.load_files()
            
            # Initialize header row variables
            self.target_header_row = self.detect_header_row(self.target_ws)
            logger.debug(f"Target header row: {self.target_header_row}")
            
            # NEW: Detect header rows in input by matching target columns
            input_header_rows = self.detect_header_rows_by_target_match()
            
            # Split input sheet into sections
            self.sections = self.split_input_sheet_by_headers(input_header_rows)
            
            if len(self.sections) == 0:
                raise Exception("No valid sections found in input file")
            
            logger.info(f"Found {len(self.sections)} sections to process")
            
            # Process each section
            output_files = []
            for i, (header_row, start_row, end_row) in enumerate(self.sections, 1):
                try:
                    section_output = self.process_sheet_section(i, header_row, start_row, end_row)
                    output_files.append(section_output)
                    safe_print(f"✓ Section {i} processed: {os.path.basename(section_output)}")
                except Exception as e:
                    logger.error(f"Failed to process section {i}: {e}")
                    safe_print(f"❌ Error processing section {i}: {e}")
            
            if not output_files:
                raise Exception("No sections were processed successfully")
            
            #Check and remove empty files
            valid_output_files = self.check_and_remove_empty_files(output_files)
            
            if not valid_output_files:
                raise Exception("All output files were empty and removed")

            # Summary
            if len(output_files) == 1:
                logger.info(f"Successfully formatted single section to: {output_files[0]}")
                safe_print(f"Successfully formatted and saved to: {output_files[0]}")
            else:
                logger.info(f"Successfully formatted {len(output_files)} sections")
                safe_print(f"Successfully formatted {len(output_files)} sections:")
                for output_file in output_files:
                    safe_print(f"  - {os.path.basename(output_file)}")
            
        except Exception as e:
            logger.error(f"Formatting failed: {e}")
            if self.error_messages:
                # Copy to problematic directory on error
                problematic_dir = Path("data/problematic")
                problematic_dir.mkdir(parents=True, exist_ok=True)
                problematic_path = problematic_dir / os.path.basename(self.input_file)
                shutil.copy2(self.input_file, problematic_path)
                logger.info(f"Copied problematic file to: {problematic_path}")
            raise e
        finally:
            # Clean up
            if self.input_wb:
                self.input_wb.close()
            if hasattr(self, 'target_wb') and self.target_wb:
                self.target_wb.close()


def format_single_file(input_file: str, target_file: str, mapping_file: str, 
                      output_file: str, input_sheet: str, target_sheet: str, 
                      table_end_tolerance: int = 1, clean_formula_only_rows: bool = True):
    formatter = ExcelFormatter(input_file, target_file, mapping_file, 
                              output_file, input_sheet, target_sheet, 
                              table_end_tolerance, clean_formula_only_rows)
    formatter.format_excel()


def format_all_sheets(input_file: str, target_file: str, mapping_file: str, target_sheet: str, 
                     table_end_tolerance: int = 1, clean_formula_only_rows: bool = True):
    """Format all sheets in an Excel file"""
    # Validate input files
    for file_path in [input_file, target_file, mapping_file]:
        if not os.path.exists(file_path):
            raise Exception(f"File not found: {file_path}")
    
    # Create results directory
    results_dir = Path("data/results")
    results_dir.mkdir(parents=True, exist_ok=True)
    
    # Get all sheet names from input file
    input_wb = load_workbook(input_file, read_only=True)
    input_sheets = input_wb.sheetnames
    input_wb.close()
    
    if not input_sheets:
        raise Exception("No sheets found in input file")
    
    logger.info(f"Processing {len(input_sheets)} sheets from {os.path.basename(input_file)}")
    
    # Process each sheet
    input_filename = Path(input_file).stem
    
    for sheet_name in input_sheets:
        output_filename = f"{input_filename}-{sheet_name}.xlsx"
        output_file = results_dir / output_filename
        
        try:
            logger.info(f"Processing sheet: {sheet_name}")
            format_single_file(
                input_file, target_file, mapping_file,
                str(output_file), sheet_name, target_sheet, 
                table_end_tolerance, clean_formula_only_rows
            )
            safe_print(f"✓ Format successful for sheet: {sheet_name}")
        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {e}")
            safe_print(f"❌ Error for sheet {sheet_name}: {e}")
            safe_print(f"Problematic file copied to: data/problematic/{os.path.basename(input_file)}")

def main():
    """Main entry point for command line usage"""
    if len(sys.argv) < 5:
        safe_print("Usage: python format_excel.py <input_file> <target_file> <mapping_file> <target_sheet> [table_end_tolerance] [clean_formula_only_rows] [output_file] [input_sheet]")
        safe_print("  If output_file is not provided, will format all sheets")
        sys.exit(1)
    
    input_file = sys.argv[1]
    target_file = sys.argv[2]
    mapping_file = sys.argv[3]
    target_sheet = sys.argv[4]
    
    # Parse table_end_tolerance parameter
    table_end_tolerance = 1  # default
    if len(sys.argv) >= 6:
        try:
            table_end_tolerance = int(sys.argv[5])
        except ValueError:
            safe_print(f"Invalid table_end_tolerance value: {sys.argv[5]}, using default 1")
    
    # Parse clean_formula_only_rows parameter
    clean_formula_only_rows = True  # default
    if len(sys.argv) >= 7:
        clean_formula_only_rows = sys.argv[6].lower() == "true"
    
    if len(sys.argv) >= 9:
        # Single file format
        output_file = sys.argv[7]
        input_sheet = sys.argv[8]
        format_single_file(input_file, target_file, mapping_file, 
                          output_file, input_sheet, target_sheet, 
                          table_end_tolerance, clean_formula_only_rows)
    else:
        # Format all sheets
        format_all_sheets(input_file, target_file, mapping_file, target_sheet, 
                         table_end_tolerance, clean_formula_only_rows)


if __name__ == "__main__":
    main()