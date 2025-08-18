#!/usr/bin/env python3
"""
Excel Reader
Reads formatted Excel files and extracts data for CSV export.
"""

import os
import re
import json
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from csv_config import CSVConfig
from picklist_handler import PicklistHandler

logger = logging.getLogger(__name__)

class ExcelReader:
    """Reads formatted Excel files and extracts SalesPlans and LineItems data"""
    
    def __init__(self, config: CSVConfig, picklist_handler: PicklistHandler):
        self.config = config
        self.picklist_handler = picklist_handler
        self.column_mappings = self.load_existing_column_mappings()
    
    def load_existing_column_mappings(self) -> Dict[str, str]:
        """
        Load existing column mappings from column_mapping.json
        
        Returns:
            Dict mapping scanned columns to target columns
        """
        mapping_file = Path("data/output/column_mapping.json")
        
        if not mapping_file.exists():
            logger.warning(f"Column mapping file not found: {mapping_file}")
            return {}
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mapping_data = json.load(f)
            
            # Convert to scanned -> target mapping
            mappings = {}
            for mapping in mapping_data.get('mappings', []):
                if not mapping.get('is_ignored', False) and mapping.get('target_column'):
                    scanned_col = mapping['scanned_column']
                    target_col = mapping['target_column']
                    mappings[target_col] = scanned_col
            
            logger.info(f"Loaded {len(mappings)} column mappings")
            return mappings
            
        except Exception as e:
            logger.error(f"Failed to load column mappings: {e}")
            return {}
    
    def extract_salesplan_metadata(self, excel_file_path: str) -> Dict[str, Any]:
        """
        Extract SalesPlans metadata from an Excel file using configurable columns
        
        Args:
            excel_file_path: Path to the Excel file
            
        Returns:
            Dict containing SalesPlans row data
        """
        filename = os.path.basename(excel_file_path)
        logger.debug(f"Extracting SalesPlans metadata from: {filename}")
        
        # Get SalesPlans column definitions
        salesplans_columns = self.config.get_salesplans_columns()
        
        # Initialize data dictionary
        salesplan_data = {}
        
        # Process each configured column
        for column_def in salesplans_columns:
            column_name = column_def["name"]
            column_type = column_def["type"]
            
            if column_type == "auto_id":
                salesplan_data[column_name] = self.config.get_next_id("salesplan")
            elif column_type == "text":
                if column_name == "SourceFile":
                    salesplan_data[column_name] = filename
                else:
                    salesplan_data[column_name] = ""
            elif column_type == "picklist":
                if column_name == "ActivityType":
                    salesplan_data[column_name] = self.picklist_handler.derive_activity_type_from_filename(filename)
                elif column_name == "Customer":
                    salesplan_data[column_name] = self.picklist_handler.apply_picklist_mapping("", column_name, "salesplans")
                else:
                    salesplan_data[column_name] = self.picklist_handler.apply_picklist_mapping("", column_name, "salesplans")
            elif column_type == "date":
                if column_name in ["StartDate", "EndDate"]:
                    # Read dates from Excel columns
                    date_value = self.extract_date_from_excel(excel_file_path, column_name)
                    salesplan_data[column_name] = date_value
                else:
                    salesplan_data[column_name] = datetime.now().strftime("%Y-%m-%d")
            elif column_type == "number":
                if column_name == "ListPriceMonth":
                    # Calculate from StartDate
                    start_date = self.extract_date_from_excel(excel_file_path, "StartDate")
                    salesplan_data[column_name] = self.extract_month_from_date(start_date)
                elif column_name == "ListPriceYear":
                    # Calculate from StartDate
                    start_date = self.extract_date_from_excel(excel_file_path, "StartDate")
                    salesplan_data[column_name] = self.extract_year_from_date(start_date)
                else:
                    salesplan_data[column_name] = 0
            else:
                salesplan_data[column_name] = ""
        
        logger.info(f"Generated SalesPlans metadata: {salesplan_data.get('SalesPlanID', 'Unknown')} - "
                   f"{salesplan_data.get('ActivityType', 'Unknown')} - "
                   f"{salesplan_data.get('StartDate', 'Unknown')} to {salesplan_data.get('EndDate', 'Unknown')}")
        return salesplan_data
    
    def extract_lineitems_data(self, excel_file_path: str, salesplan_id: str) -> List[Dict[str, Any]]:
        """
        Extract LineItems data from an Excel file
        
        Args:
            excel_file_path: Path to the Excel file
            salesplan_id: Associated SalesPlans ID
            
        Returns:
            List of dicts containing LineItems data
        """
        filename = os.path.basename(excel_file_path)
        logger.debug(f"Extracting LineItems data from: {filename}")
        
        try:
            # Load Excel file
            wb = load_workbook(excel_file_path, data_only=True)
            ws = wb.active
            
            # Detect header row
            header_row = self.detect_header_row(ws)
            
            # Get headers mapping
            headers = self.get_headers_mapping(ws, header_row)
            
            # Get source mapping for LineItems columns
            source_mapping = self.config.get_column_source_mapping("lineitems")
            
            # Derive product category from filename
            product_category = self.picklist_handler.derive_product_category_from_filename(filename)
            
            lineitems = []
            
            # Process each data row
            for row_idx in range(header_row + 1, ws.max_row + 1):
                lineitem_data = self.extract_single_lineitem(
                    ws, row_idx, headers, source_mapping, salesplan_id, product_category
                )
                
                if lineitem_data:  # Only add if we got valid data
                    lineitems.append(lineitem_data)
            
            wb.close()
            logger.info(f"Extracted {len(lineitems)} line items from {filename}")
            return lineitems
            
        except Exception as e:
            logger.error(f"Failed to extract LineItems from {filename}: {e}")
            return []
    
    def extract_single_lineitem(self, ws, row_idx: int, headers: Dict[str, int], 
                               source_mapping: Dict[str, str], salesplan_id: str, 
                               product_category: str) -> Optional[Dict[str, Any]]:
        """
        Extract a single LineItem from a row
        
        Args:
            ws: Worksheet object
            row_idx: Row index to process
            headers: Headers mapping
            source_mapping: CSV column to source column mapping
            salesplan_id: Associated SalesPlans ID
            product_category: Product category for this file
            
        Returns:
            LineItem data dict or None if row is empty
        """
        lineitem_data = {}
        has_data = False
        
        # Generate LineItem ID
        lineitem_data["LineItemID"] = self.config.get_next_id("lineitem")
        lineitem_data["SalesPlanID"] = salesplan_id
        lineitem_data["ProductCategory"] = product_category
        
        # Process each CSV column
        for csv_column, source_column in source_mapping.items():
            if csv_column in ["LineItemID", "SalesPlanID", "ProductCategory"]:
                continue  # Already handled above
            
            # Find the source column in headers
            cell_value = None
            if source_column in headers:
                col_idx = headers[source_column]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # Check if we have any non-empty data
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
            
            # Apply picklist mapping if needed
            if self.config.is_picklist_column(csv_column, "lineitems"):
                mapped_value = self.picklist_handler.apply_picklist_mapping(
                    cell_value, csv_column, "lineitems"
                )
                lineitem_data[csv_column] = mapped_value
            else:
                # Handle different data types
                lineitem_data[csv_column] = self.format_cell_value(cell_value, csv_column)
        
        # Only return if we have meaningful data
        if not has_data:
            return None
        
        logger.debug(f"Extracted LineItem: {lineitem_data.get('ProductBarcode', 'No barcode')}")
        return lineitem_data
    
    def detect_header_row(self, ws) -> int:
        """
        Detect header row in worksheet
        
        Args:
            ws: Worksheet object
            
        Returns:
            Header row number (1-based)
        """
        # Use same logic as existing format_excel.py
        if ws.max_row == 0:
            return 1
        
        # Find the rightmost column with data
        rightmost_col = 0
        for row_idx in range(1, min(ws.max_row + 1, 20)):  # Check first 20 rows
            for col_idx in range(ws.max_column, 0, -1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    if col_idx > rightmost_col:
                        rightmost_col = col_idx
                    break
        
        if rightmost_col == 0:
            return 1
        
        # Find first row with data in rightmost column
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=rightmost_col)
            if cell.value is not None and str(cell.value).strip():
                return row_idx
        
        return 1
    
    def get_headers_mapping(self, ws, header_row: int) -> Dict[str, int]:
        """
        Get mapping of header names to column indices
        
        Args:
            ws: Worksheet object
            header_row: Header row number
            
        Returns:
            Dict mapping header name to column index
        """
        headers = {}
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value:
                header_name = str(cell.value).strip()
                if header_name:
                    headers[header_name] = col_idx
        
        logger.debug(f"Found {len(headers)} headers in row {header_row}")
        return headers
    
    def extract_date_from_excel(self, excel_file_path: str, date_field: str) -> str:
        """
        Extract date from Excel columns using configurable column names
        
        Args:
            excel_file_path: Path to Excel file
            date_field: "StartDate" or "EndDate"
            
        Returns:
            Date string in YYYY-MM-DD format
        """
        try:
            # Get Excel column mappings
            excel_mappings = self.config.get_excel_column_mappings()
            
            if date_field == "StartDate":
                column_name = excel_mappings.get("start_date_column", "Başlangıç")
            elif date_field == "EndDate":
                column_name = excel_mappings.get("end_date_column", "Bitiş")
            else:
                logger.warning(f"Unknown date field: {date_field}")
                return datetime.now().strftime("%Y-%m-%d")
            
            # Load Excel file
            wb = load_workbook(excel_file_path, data_only=True)
            ws = wb.active
            
            # Detect header row
            header_row = self.detect_header_row(ws)
            
            # Get headers mapping
            headers = self.get_headers_mapping(ws, header_row)
            
            # Find the date column
            if column_name not in headers:
                logger.warning(f"Date column '{column_name}' not found in Excel file {excel_file_path}")
                wb.close()
                return datetime.now().strftime("%Y-%m-%d")
            
            col_idx = headers[column_name]
            
            # Read date value from first data row
            for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    date_value = self.format_cell_value(cell.value, date_field)
                    wb.close()
                    logger.debug(f"Extracted {date_field} from column '{column_name}': {date_value}")
                    return date_value
            
            wb.close()
            logger.warning(f"No date value found in column '{column_name}' for {date_field}")
            return datetime.now().strftime("%Y-%m-%d")
            
        except Exception as e:
            logger.error(f"Failed to extract {date_field} from Excel file {excel_file_path}: {e}")
            return datetime.now().strftime("%Y-%m-%d")
    
    def extract_month_from_date(self, date_str: str) -> int:
        """Extract month from date string"""
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj.month
        except:
            return datetime.now().month
    
    def extract_year_from_date(self, date_str: str) -> int:
        """Extract year from date string"""
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj.year
        except:
            return datetime.now().year
    
    def format_cell_value(self, cell_value: Any, column_name: str) -> str:
        """
        Format cell value based on column type
        
        Args:
            cell_value: Raw cell value
            column_name: Name of the column
            
        Returns:
            Formatted string value
        """
        if cell_value is None:
            return ""
        
        # Handle dates
        if isinstance(cell_value, (date, datetime)):
            return cell_value.strftime("%Y-%m-%d")
        
        # Handle numbers
        if isinstance(cell_value, (int, float)):
            if column_name.endswith("Date"):
                # Excel date serial number
                try:
                    excel_date = datetime(1900, 1, 1) + timedelta(days=cell_value - 2)
                    return excel_date.strftime("%Y-%m-%d")
                except:
                    return str(cell_value)
            else:
                return str(cell_value)
        
        # Default: convert to string
        return str(cell_value).strip()